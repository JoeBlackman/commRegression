'''
************************************************************************************************************************
*File       regressionDirector.py
*Author     Joseph Blackman
*Date       5/21/2018
*Brief      Python 3.6 command line script for running automated communication regression tests.
*           A module of the comm regression scripts package (ASCO P/N: 1258991)
*Source file Revision History
*
*Version 1.00
*   5/21/2018 JBLACKMAN
*       1. Initial Release
************************************************************************************************************************
'''

#built in modules
import argparse
import datetime
import sys
from sqlalchemy import exc

#local modules
import regressionTest
import utilities
#from ascoparser import parser
from statusLogger import status_logger
from resultsLogger import results_logger
import connection
import test

from databaseHandler import dataAccessObject as dao
from databaseHandler import dataTransferObject as dto

#3rd party modules
from openpyxl import load_workbook

def parseArguments():
    parser = argparse.ArgumentParser(description='''ASCO P/N: 1258991 \r\n 
    A command line script for automated communications testing.''', formatter_class=argparse.RawTextHelpFormatter)
    parser.add_argument('-i', '--inputCDS', nargs=1, required=True, help='path to CDS')
    parser.add_argument('-d', '--device_name', nargs=1, required=True, help='Device name as specified by sheet name in cds ')
    parser.add_argument('-c', '--connections', nargs='+', help='''
    Definitions for each connection to a device. The arguments are formatted differently depending on the type
    of connection. For multiple connections, separate with a space.
    CAN:
    'canbus',protocol,baud_rate,node_id
    example: canbus,canopensdo,500000,0x1C
    Ethernet:
    'ethernet',protocol,ip_address,tcp_port,unit_id_tcp
    example: ethernet,modbus,169.254.1.1,502,1
    Serial:
    'serial',protocol,port,baud_rate,byte_size,parity,stop_bits,xonxoff,unit_id_serial,timeout
    example: serial,modbus,COM1,9600,8,N,1,0,1,5''')
    parser.add_argument('-f', '--filters', nargs='+', help='''filters what tests will be performed. The arguments are 
    formatted as follows:
    filterName1:filterValue1 filterName2:filterValue2
    where filterName is the name of the filter you wish to apply and filterValue is the value you wish to filter by.
    Legal Filter Names:
    device_name,
    node_id,
    can_baud_rate,
    ip_address,
    tcp_port,
    serial_port,
    unit_id_tcp,
    unit_id_rtu,
    modbus_baud_rate,
    test_type,
    protocol,
    can_object,
    modbus_register,
    user,
    access_for_tcp,
    access_for_can,
    access_for_ttl,
    access_for_rs485,
    parameter_name,
    parameter_description,
    data_type,
    length,
    data_range_min,
    data_range_max,
    data_range_multiplier,
    unit,
    supported_firmware_version,
    note''')
    args = parser.parse_args()
    return args

def setupLogging(statusLogPath = None, resultsLogPath = None):
    # clears current log files
    if statusLogPath is not None:
        with open(statusLogPath, 'w'):
            pass
    else:
        with open('status.log', 'w'):
            pass
    if resultsLogPath is not None:
        with open(resultsLogPath, 'w'):
            pass
    else:
        with open('results.csv', 'w'):
            pass
    status_logger.info(datetime.date.today())

def buildSchema():
    testsDataAccessObject = dao.testDAO()
    connectionDataAccessObject = dao.connectionDAO()
    try:
        testsDataAccessObject.drop()
    except exc.NoSuchTableError:
        pass
    testsDataAccessObject.create()
    try:
        connectionDataAccessObject.drop()
    except exc.NoSuchTableError:
        pass
    connectionDataAccessObject.create()

#def getTableHeaders(inputArgs):
#    #Create a list of what arguments were used ['inputCDS', 'device_name', 'ip_address', 'tcp_port', 'unit_id_tcp']
#    inputHeaders = []
#    for k, v in (vars(inputArgs)).items():
#        if (v is not None) & (k != 'inputCDS'):
#            inputHeaders.append(k)
#    #read cds into memory and store it's headers
#    [cds] = inputArgs.inputCDS
#    wb = load_workbook(cds)
#    ws = wb[inputArgs.device_name[0]]
#    #TODO: extrapolate ranges no longer works
#    data = utilities.cellsToValues(ws)
#    data = utilities.removeExtraColumns(data)
#    data = utilities.extrapolateRanges(data)
#    cdsHeaders = data[0]
#    inputHeaders.extend(cdsHeaders)
#    #concatenate input args headers to cds headers
#    return inputHeaders

def createTests(args):
    #assign command line arguments related to connections to location in database table connections
    connDataAccessObject = dao.connectionDAO()
    connDict = {}
    connectionCount = 0
    for item in args.connections:
        connectionCount = connectionCount + 1
        transport = None
        protocol = None
        canbaudrate = None
        nodeid = None
        ipaddress = None
        tcpport = None
        tcpunitid = None
        serialport = None
        serialbaudrate = None
        serialbytesize = None
        serialparity = None
        serialstopbits = None
        serialxonxoff = None
        serialunitid = None
        serialtimeout = None
        split = item.split(',')
        if split[0] == 'canbus':
            transport = split[0]
            protocol = split[1]
            canbaudrate = split[2]
            nodeid = split[3]
        elif split[0] == 'ethernet':
            transport = split[0]
            protocol = split[1]
            ipaddress = split[2]
            tcpport = split[3]
            tcpunitid = split[4]
        elif split[0] == 'serial':
            transport = split[0]
            protocol = split[1]
            serialport = split[2]
            serialbaudrate = split[3]
            serialbytesize = split[4]
            serialparity = split[5]
            serialstopbits = split[6]
            serialxonxoff = split[7]
            serialunitid = split[8]
            serialtimeout = split[9]
        c = connection.connection(can_node_id=nodeid, can_baud_rate=canbaudrate, ip_address=ipaddress, tcp_port=tcpport,
                                  tcp_unit_id=tcpunitid, serial_port=serialport, serial_baud_rate=serialbaudrate,
                                  serial_byte_size=serialbytesize, serial_parity=serialparity,
                                  serial_stop_bits=serialstopbits, serial_xonxoff=serialxonxoff,
                                  serial_unit_id=serialunitid, serial_timeout=serialtimeout, protocol=protocol,
                                  transport=transport,)
        connDataAccessObject.insert(c)
        connDict[str(connectionCount)] = c
    for k,v in connDict.items():
        print(v.protocol, v.transport)

    #get the communication object data from the cds
    [cds] = args.inputCDS
    wb = load_workbook(cds)
    ws = wb[args.device_name[0]]
    data = utilities.cellsToValues(ws)
    data = utilities.removeExtraColumns(data)
    data = utilities.extrapolateRanges(data)

    headers = data[0]
    #for each row in the cds, construct each possible test and add the test to a master list
    possibleTests = []
    for row in data:
        d = dict(zip(headers, row))
        #print(d)
        #look for canbus_canopensdo_access column
        #if it isn't None, assume protocol = sdo, transport = canbus
        if d['canbus_canopen_access'] is not None:
            canConnections = []
            for k,v in connDict.items():
                if (v.protocol == 'canopensdo') & (v.transport == 'canbus'):
                    #just create the test here


        #find connections with matching protocol and transport
        #for each matching connection, create a single read test and a write to read only test and append them to
        #list of possible tests
        #look for ethernet_modbus_access column
        #print(d['ethernet_modbus_tcp_access'])
        #look for rs485_modbus_access column

def createTest(testInfo, deviceInfo): #tuple, tuple: returns test object?
    pass

def buildRegressionTest(database):
    status_logger.info('Building Regression Test')
    try:
        rt = regressionTest.regressionTest(database)
        rt.executeTests()
        cleanup(database)
    #except AssertionError:
    #    status_logger.info('Number of sheets, node Ids, and CAN Baud Rates must match')
    #    cleanup(database)
    except KeyboardInterrupt:
        status_logger.info('Keyboard Interrupt Detected')
        cleanup(database)
    except Exception as e:
        status_logger.info('something went wrong... %s' % repr(e))
        cleanup(database)

def cleanup(database):
    database.close()
    status_logger.info('End Regression')

def main(argv):
    # Start of script execution
    args = parseArguments()
    buildSchema()
    print(args)
    createTests(args)

    #setupLogging()
    #just got code to work up to this point with new architecture (easier command line, modbustcp configurable)
    #buildRegressionTest(database)

if __name__ == "__main__":
    main(sys.argv[1:])
